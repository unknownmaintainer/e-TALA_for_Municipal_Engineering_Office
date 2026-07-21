import logging
import os
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden, FileResponse
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, F
from django.utils import timezone
from django.core import signing
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.sessions.models import Session

from .models import (
    CustomUser, Barangay, Category, Record, Document,
    AuditLog, LoginAttempt, PasswordHistory,
    EngineeringRecord, PermitDetail, ProjectDetail,
    RequirementTemplate, RequirementItem, RecordRequirement,
    BlockedIP,
)
from .validators import validate_document_file, sanitize_input
from .utils import get_client_ip

logger = logging.getLogger('permits')


def log_audit(user, action, target_record_id=None, request=None):
    # Prevent logging repetitive routine activities to keep AuditLogs clean and readable
    ignored_actions = [
        "Logged out",
        "Logged in successfully",
        "Logged in successfully from a new IP/device"
    ]
    if action in ignored_actions:
        return
        
    ip = get_client_ip(request) if request else None
    AuditLog.objects.create(
        user=user,
        action=action,
        target_record_id=target_record_id,
        ip_address=ip
    )


def check_lockout(email, ip_address):
    fifteen_mins_ago = timezone.now() - timedelta(minutes=15)
    total_failures = LoginAttempt.objects.filter(email_attempted=email, success=False).count()
    if total_failures >= 10:
        User = get_user_model()
        user_obj = User.objects.filter(email=email).first()
        if user_obj and user_obj.is_active:
            user_obj.is_active = False
            user_obj.save()
            log_audit(user_obj, "Account locked out permanently (10 failed attempts)", request=None)
        return True, "Account locked. Please contact the administrator for unlock."

    recent_failures = LoginAttempt.objects.filter(
        email_attempted=email,
        success=False,
        timestamp__gte=fifteen_mins_ago
    ).order_by('-timestamp')

    if recent_failures.count() >= 5:
        fifth_failure = recent_failures[4]
        elapsed = timezone.now() - fifth_failure.timestamp
        remaining = 15 - int(elapsed.total_seconds() / 60)
        if remaining > 0:
            return True, f"Account temporarily locked. Try again in {remaining} minutes."

    return False, None


def validate_password_strength(password):
    if len(password) < 8:
        return False, "Password must be at least 8 characters long."
    if not any(c.isupper() for c in password):
        return False, "Password must contain at least one uppercase letter."
    if not any(c.islower() for c in password):
        return False, "Password must contain at least one lowercase letter."
    if not any(c.isdigit() for c in password):
        return False, "Password must contain at least one number."
    return True, None


def get_per_page(request, default=10):
    val = request.GET.get('per_page', '')
    try:
        val = int(val)
        if val in [10, 20, 50, 100]:
            return val
    except ValueError:
        pass
    return default


def get_year_choices():
    db_years = set(EngineeringRecord.objects.exclude(year__isnull=True).values_list('year', flat=True))
    from datetime import date
    current_year = date.today().year
    default_years = set(range(current_year + 1, current_year - 15, -1))
    return sorted(list(db_years.union(default_years)), reverse=True)



# ─── PUBLIC LANDING ─────────────────────────────────────────────────────────

def landing_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


# ─── AUTHENTICATION ──────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '')
        ip_address = get_client_ip(request)

        is_locked, lockout_msg = check_lockout(email, ip_address)
        if is_locked:
            messages.error(request, lockout_msg)
            return render(request, 'permits/login.html')

        User = get_user_model()
        user_obj = User.objects.filter(email=email).first()
        username = user_obj.username if user_obj else None

        user = authenticate(request, username=username, password=password)
        if user is not None:
            if not user.is_active:
                LoginAttempt.objects.create(email_attempted=email, success=False, ip_address=ip_address)
                messages.error(request, "Account is locked. Please contact the administrator.")
                return render(request, 'permits/login.html')

            if user.session_key:
                try:
                    Session.objects.filter(session_key=user.session_key).delete()
                except Exception as e:
                    logger.error(f"Error terminating previous session: {e}")

            LoginAttempt.objects.create(email_attempted=email, success=True, ip_address=ip_address)
            login(request, user)

            user.session_key = request.session.session_key
            user.save()

            # Check if this is a login from a new IP/device
            has_previous_login = AuditLog.objects.filter(
                user=user,
                action__icontains="Logged in",
                ip_address=ip_address
            ).exists()
            
            action_text = "Logged in successfully"
            if ip_address and not has_previous_login:
                action_text = "Logged in successfully from a new IP/device"

            log_audit(user, action_text, request=request)
            return redirect('dashboard')
        else:
            LoginAttempt.objects.create(email_attempted=email, success=False, ip_address=ip_address)
            if user_obj:
                log_audit(user_obj, "Failed login attempt", request=request)
            messages.error(request, "Incorrect email or password.")

    return render(request, 'permits/login.html')


def logout_view(request):
    if request.user.is_authenticated:
        log_audit(request.user, "Logged out", request=request)
        request.user.session_key = None
        request.user.save()
    logout(request)
    return redirect('login')


def forgot_password_view(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        messages.info(request, "If the email is registered, a password reset link has been sent.")
        return redirect('login')
    return render(request, 'permits/forgot_password.html')


def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        full_name = sanitize_input(request.POST.get('full_name', '')).strip()
        email = sanitize_input(request.POST.get('email', '')).strip().lower()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        government_id = request.FILES.get('government_id')

        if not full_name or not email or not password or not confirm_password or not government_id:
            messages.error(request, "All required fields must be filled, including your Government ID.")
            return render(request, 'permits/register.html')

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'permits/register.html')

        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, "An account with this email already exists.")
            return render(request, 'permits/register.html')

        username = email.split('@')[0]
        base_username = username
        counter = 1
        while CustomUser.objects.filter(username=username).exists():
            username = f"{base_username}{counter}"
            counter += 1

        ok, err_msg = validate_password_strength(password)
        if not ok:
            messages.error(request, err_msg)
            return render(request, 'permits/register.html')

        User = get_user_model()
        new_user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            full_name=full_name,
            role='staff',
            profile_picture=government_id
        )
        new_user.save()

        log_audit(new_user, "Registered account", request=request)
        messages.success(request, "Account created successfully! Please sign in.")
        return redirect('login')

    return render(request, 'permits/register.html')


# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@login_required
def dashboard_view(request):
    records = EngineeringRecord.objects.exclude(status='archived')

    # Summary stats
    total_permits = records.filter(record_type='Permit').count()
    total_municipal = records.filter(record_type='Project', project_scope='Municipal').count()
    total_barangay = records.filter(record_type='Project', project_scope='Barangay').count()
    total_documents = Document.objects.filter(engineering_record__isnull=False).count()
    total_archived = EngineeringRecord.objects.filter(status='archived').count()
    total_records = records.count()

    # Incomplete records (have a checklist but not fully fulfilled or waived)
    incomplete_records = records.filter(
        requirements__isnull=False,
        requirements__is_fulfilled=False,
        requirements__is_waived=False
    ).distinct().count()

    # Checklist Digitization Compliance Stats
    active_with_reqs = records.annotate(
        total_reqs=Count('requirements'),
        fulfilled_reqs=Count('requirements', filter=Q(requirements__is_fulfilled=True) | Q(requirements__is_waived=True))
    ).filter(total_reqs__gt=0)
    
    compliance_total = active_with_reqs.count()
    compliance_completed = active_with_reqs.filter(total_reqs=F('fulfilled_reqs')).count()
    compliance_incomplete = compliance_total - compliance_completed
    compliance_rate = int((compliance_completed / compliance_total) * 100) if compliance_total > 0 else 100

    # Recent records
    recent_records = records.select_related(
        'barangay', 'created_by', 'permit_detail', 'project_detail'
    ).order_by('-created_at')[:8]

    # Activity feed
    if request.user.role == 'admin':
        activity_feed = AuditLog.objects.select_related('user').order_by('-performed_at')[:15]
    else:
        activity_feed = AuditLog.objects.filter(user=request.user).select_related('user').order_by('-performed_at')[:15]

    # Recent Uploads
    recent_uploads = Document.objects.select_related('engineering_record', 'uploaded_by').order_by('-uploaded_at')[:5]

    # Incomplete Records list
    incomplete_list = records.filter(
        requirements__isnull=False,
        requirements__is_fulfilled=False,
        requirements__is_waived=False
    ).distinct().select_related('barangay')[:15]

    # Pending records
    pending_records = records.filter(status='pending').select_related('barangay').order_by('-created_at')[:5]

    # Warning alerts context data (grouped, optimized for scalability with preview & count)
    failed_logins_count = 0
    if request.user.role == 'admin':
        failed_logins_count = LoginAttempt.objects.filter(success=False).count()
        
    today_date = timezone.now().date()
    thirty_days_later = today_date + timedelta(days=30)
    
    alert_docs = Document.objects.filter(
        expiry_date__isnull=False
    ).exclude(engineering_record__status='archived').select_related('engineering_record', 'requirement_item')
    
    expired_docs_qs = alert_docs.filter(expiry_date__lt=today_date)
    expired_count = expired_docs_qs.count()
    expired_preview = []
    for doc in expired_docs_qs.order_by('-expiry_date')[:3]:
        doc_label = doc.requirement_item.name if doc.requirement_item else doc.document_type
        expired_preview.append({
            'label': doc_label,
            'record_title': doc.engineering_record.title,
            'record_id': doc.engineering_record.record_id,
            'expiry_date': doc.expiry_date.strftime('%b %d, %Y'),
        })
        
    expiring_docs_qs = alert_docs.filter(expiry_date__range=(today_date, thirty_days_later))
    expiring_count = expiring_docs_qs.count()
    expiring_preview = []
    for doc in expiring_docs_qs.order_by('expiry_date')[:3]:
        doc_label = doc.requirement_item.name if doc.requirement_item else doc.document_type
        expiring_preview.append({
            'label': doc_label,
            'record_title': doc.engineering_record.title,
            'record_id': doc.engineering_record.record_id,
            'expiry_date': doc.expiry_date.strftime('%b %d, %Y'),
        })

    alerts = {
        'failed_logins': failed_logins_count,
        'expired_count': expired_count,
        'expired_preview': expired_preview,
        'expiring_count': expiring_count,
        'expiring_preview': expiring_preview,
        'has_alerts': (failed_logins_count > 0 or expired_count > 0 or expiring_count > 0)
    }

    # Charts data
    import json
    # 1. Records per Year (last 5 years)
    years_query = records.filter(year__isnull=False).values('year').annotate(count=Count('record_id')).order_by('year')[:5]
    chart_years = [str(y['year']) for y in years_query]
    chart_year_counts = [y['count'] for y in years_query]

    # 2. Permit Distribution
    permit_query = PermitDetail.objects.values('permit_type').annotate(count=Count('id')).order_by('-count')
    chart_permits = [p['permit_type'] for p in permit_query]
    chart_permit_counts = [p['count'] for p in permit_query]

    # Illegal Construction Tracking Stats
    illegal_qs = records.filter(is_illegal_construction=True)
    illegal_total = illegal_qs.count()
    illegal_unresolved = illegal_qs.filter(illegal_compliance_status='unresolved').count()
    illegal_pending = illegal_qs.filter(illegal_compliance_status='pending_permit').count()
    illegal_resolved = illegal_qs.filter(illegal_compliance_status='resolved').count()
    illegal_recent = illegal_qs.select_related('barangay').order_by('-created_at')[:5]

    context = {
        'total_permits': total_permits,
        'total_municipal': total_municipal,
        'total_barangay': total_barangay,
        'total_documents': total_documents,
        'total_archived': total_archived,
        'total_records': total_records,
        'incomplete_records': incomplete_records,
        'recent_records': recent_records,
        'activity_feed': activity_feed,
        'recent_uploads': recent_uploads,
        'incomplete_list': incomplete_list,
        'pending_records': pending_records,
        'illegal_total': illegal_total,
        'illegal_unresolved': illegal_unresolved,
        'illegal_pending': illegal_pending,
        'illegal_resolved': illegal_resolved,
        'illegal_recent': illegal_recent,
        'alerts': alerts,
        'barangays': Barangay.objects.all(),
        'compliance_total': compliance_total,
        'compliance_completed': compliance_completed,
        'compliance_incomplete': compliance_incomplete,
        'compliance_rate': compliance_rate,
        'chart_years_json': json.dumps(chart_years),
        'chart_year_counts_json': json.dumps(chart_year_counts),
        'chart_permits_json': json.dumps(chart_permits),
        'chart_permit_counts_json': json.dumps(chart_permit_counts),
        'active_tab': 'dashboard',
    }
    return render(request, 'permits/dashboard.html', context)


# ─── BARANGAYS ───────────────────────────────────────────────────────────────

@login_required
def barangays_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if request.user.role not in ['admin', 'staff']:
            return HttpResponseForbidden("Unauthorized")

        if action == 'add_barangay':
            name = request.POST.get('barangay_name', '').strip()
            if name:
                if Barangay.objects.filter(barangay_name__iexact=name).exists():
                    messages.error(request, f"Barangay '{name}' already exists.")
                else:
                    b = Barangay.objects.create(barangay_name=name)
                    log_audit(request.user, f"Created Barangay '{b.barangay_name}'", request=request)
                    messages.success(request, f"Barangay '{b.barangay_name}' has been created successfully.")
            return redirect('barangays')

        elif action == 'edit_barangay':
            barangay_id = request.POST.get('barangay_id')
            name = request.POST.get('barangay_name', '').strip()
            if barangay_id and name:
                barangay = get_object_or_404(Barangay, barangay_id=barangay_id)
                if Barangay.objects.filter(barangay_name__iexact=name).exclude(barangay_id=barangay_id).exists():
                    messages.error(request, f"Another barangay named '{name}' already exists.")
                else:
                    old_name = barangay.barangay_name
                    barangay.barangay_name = name
                    barangay.save()
                    log_audit(request.user, f"Renamed Barangay '{old_name}' to '{name}'", request=request)
                    messages.success(request, f"Barangay '{old_name}' renamed to '{name}' successfully.")
            return redirect('barangays')

        elif action == 'delete_barangay':
            barangay_id = request.POST.get('barangay_id')
            if barangay_id:
                barangay = get_object_or_404(Barangay, barangay_id=barangay_id)
                records_count = barangay.engineering_records.count() + barangay.records.count()
                if records_count > 0:
                    messages.error(request, f"Cannot delete '{barangay.barangay_name}' because it contains {records_count} active record(s).")
                else:
                    name = barangay.barangay_name
                    barangay.delete()
                    log_audit(request.user, f"Deleted Barangay '{name}'", request=request)
                    messages.success(request, f"Barangay '{name}' deleted successfully.")
            return redirect('barangays')

    barangays = Barangay.objects.annotate(
        total_records=Count('engineering_records'),
        permit_count=Count('engineering_records', filter=Q(engineering_records__record_type='Permit')),
        project_count=Count('engineering_records', filter=Q(engineering_records__record_type='Project')),
    ).order_by('barangay_name')

    query = request.GET.get('q', '').strip()
    if query:
        barangays = barangays.filter(barangay_name__icontains=query)

    context = {
        'barangays': barangays,
        'q': query,
        'active_tab': 'barangays',
    }
    return render(request, 'permits/barangays.html', context)


@login_required
def barangay_workspace_view(request, barangay_id):
    barangay = get_object_or_404(Barangay, barangay_id=barangay_id)
    records = EngineeringRecord.objects.filter(barangay=barangay).exclude(status='archived').select_related('created_by')

    # Stats
    total_permits = records.filter(record_type='Permit').count()
    total_projects = records.filter(record_type='Project').count()
    total_documents = Document.objects.filter(engineering_record__barangay=barangay).count()

    # Tab filter
    tab = request.GET.get('tab', 'all')
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')
    
    filtered_records = records
    if tab == 'permits':
        filtered_records = filtered_records.filter(record_type='Permit')
    elif tab == 'projects':
        filtered_records = filtered_records.filter(record_type='Project')

    if query:
        filtered_records = filtered_records.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )
    if status_filter:
        filtered_records = filtered_records.filter(status=status_filter)

    filtered_records = filtered_records.order_by('-created_at')

    # Permits breakdown
    permit_breakdown = PermitDetail.objects.filter(
        engineering_record__barangay=barangay
    ).values('permit_type').annotate(count=Count('id')).order_by('-count')

    # Project breakdown
    project_breakdown = ProjectDetail.objects.filter(
        engineering_record__barangay=barangay
    ).values('project_type').annotate(count=Count('id')).order_by('-count')

    # Recent activity
    record_ids = records.values_list('record_id', flat=True)
    recent_activity = AuditLog.objects.filter(
        target_record_id__in=record_ids
    ).select_related('user').order_by('-performed_at')[:10]

    per_page = get_per_page(request, 10)
    paginator = Paginator(filtered_records, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'per_page': per_page,
        'barangay': barangay,
        'total_permits': total_permits,
        'total_projects': total_projects,
        'total_documents': total_documents,
        'total_records': total_permits + total_projects,
        'permit_breakdown': permit_breakdown,
        'project_breakdown': project_breakdown,
        'recent_activity': recent_activity,
        'page_obj': page_obj,
        'current_tab': tab,
        'q': query,
        'selected_status': status_filter,
        'active_tab': 'barangays',
    }
    return render(request, 'permits/barangay_workspace.html', context)


# ─── ENGINEERING RECORDS (BROWSE ALL) ────────────────────────────────────────

@login_required
def records_browse_view(request):
    base_records = EngineeringRecord.objects.exclude(status='archived').select_related('barangay', 'created_by', 'permit_detail', 'project_detail')
    barangays = Barangay.objects.all()

    # Filters
    query = request.GET.get('q', '').strip()
    record_type = request.GET.get('record_type', '')
    project_scope = request.GET.get('project_scope', '')
    barangay_id = request.GET.get('barangay', '')
    status = request.GET.get('status', '')
    year = request.GET.get('year', '')
    
    project_type = ''
    permit_type = ''
    if record_type == 'Permit':
        permit_type = request.GET.get('permit_type', '')
    else:
        project_type = request.GET.get('project_type', '')

    if query:
        search_filter = (
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(barangay__barangay_name__icontains=query) |
            Q(permit_detail__permit_number__icontains=query) |
            Q(permit_detail__applicant_name__icontains=query) |
            Q(permit_detail__permit_type__icontains=query) |
            Q(permit_detail__building_type__icontains=query) |
            Q(project_detail__project_type__icontains=query) |
            Q(project_detail__contractor__icontains=query) |
            Q(project_detail__funding_source__icontains=query) |
            Q(record_type__icontains=query) |
            Q(project_scope__icontains=query) |
            Q(illegal_compliance_status__icontains=query) |
            Q(status__icontains=query) |
            Q(created_by__full_name__icontains=query) |
            Q(created_by__username__icontains=query)
        )
        if query.isdigit():
            search_filter |= Q(year=int(query)) | Q(created_at__year=int(query)) | Q(date_started__year=int(query))
        base_records = base_records.filter(search_filter).distinct()
    
    if barangay_id:
        base_records = base_records.filter(barangay_id=barangay_id)
    if status:
        base_records = base_records.filter(status=status)
    if year:
        try:
            base_records = base_records.filter(year=int(year))
        except (ValueError, TypeError):
            base_records = base_records.filter(year=year)
    if record_type == 'Permit' and permit_type:
        base_records = base_records.filter(permit_detail__permit_type=permit_type)
    elif project_type:
        base_records = base_records.filter(project_detail__project_type=project_type)

    illegal_filter = request.GET.get('illegal', '').strip()
    if illegal_filter == '1' or illegal_filter == 'true' or record_type == 'Illegal':
        base_records = base_records.filter(is_illegal_construction=True)
    elif illegal_filter in ['unresolved', 'pending_permit', 'resolved']:
        base_records = base_records.filter(is_illegal_construction=True, illegal_compliance_status=illegal_filter)

    # Compute tab counts dynamically based on active search criteria
    all_count = base_records.count()
    municipal_count = base_records.filter(record_type='Project', project_scope='Municipal').count()
    barangay_count = base_records.filter(record_type='Project', project_scope='Barangay').count()
    permits_count = base_records.filter(record_type='Permit').count()
    illegal_count = base_records.filter(is_illegal_construction=True).count()
    illegal_unresolved_count = base_records.filter(is_illegal_construction=True, illegal_compliance_status='unresolved').count()
    illegal_pending_count = base_records.filter(is_illegal_construction=True, illegal_compliance_status='pending_permit').count()
    illegal_resolved_count = base_records.filter(is_illegal_construction=True, illegal_compliance_status='resolved').count()

    # Apply tab filter
    records = base_records
    if record_type and record_type != 'Illegal':
        records = records.filter(record_type=record_type)
    if project_scope:
        records = records.filter(project_scope=project_scope)

    total_count = records.count()
    per_page = get_per_page(request, 10)
    paginator = Paginator(records, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    year_choices = get_year_choices()

    # Calculate active advanced filters count (only count optional dropdown filters)
    active_filters_count = sum(1 for val in [barangay_id, status, year, project_type, permit_type] if val)

    context = {
        'per_page': per_page,
        'barangays': barangays,
        'page_obj': page_obj,
        'total_count': total_count,
        'all_count': all_count,
        'municipal_count': municipal_count,
        'barangay_count': barangay_count,
        'permits_count': permits_count,
        'illegal_count': illegal_count,
        'illegal_unresolved_count': illegal_unresolved_count,
        'illegal_pending_count': illegal_pending_count,
        'illegal_resolved_count': illegal_resolved_count,
        'active_filters_count': active_filters_count,
        'q': query,
        'selected_record_type': record_type,
        'selected_illegal': illegal_filter,
        'selected_project_scope': project_scope,
        'selected_barangay': barangay_id,
        'selected_status': status,
        'selected_year': year,
        'selected_project_type': project_type,
        'selected_permit_type': permit_type,
        'year_choices': year_choices,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'project_type_choices': [choice[0] for choice in ProjectDetail.PROJECT_TYPE_CHOICES],
        'permit_types': PermitDetail.PERMIT_TYPE_CHOICES,
        'active_tab': 'records',
    }
    return render(request, 'permits/records_browse.html', context)



# ─── CREATE RECORD — STEP 1: CATEGORY ────────────────────────────────────────


@login_required

@login_required
def record_create_step1_view(request):
    if request.user.role not in ['staff', 'admin']:
        raise PermissionDenied("You do not have permission to encode records.")
    
    # Pre-select category via GET parameter for quick actions
    cat_param = request.GET.get('category')
    if cat_param in ['municipal', 'barangay', 'permit']:
        request.session['create_category'] = cat_param
        return redirect('create_step2')
        
    if request.method == 'POST':
        category = request.POST.get('category')
        if category in ['municipal', 'barangay', 'permit']:
            request.session['create_category'] = category
            return redirect('create_step2')
        messages.error(request, "Invalid category selection.")
    
    return render(request, 'permits/create_step1.html', {
        'active_tab': 'records'
    })

@login_required
def record_create_step2_view(request):
    if request.user.role not in ['staff', 'admin']:
        raise PermissionDenied("You do not have permission to encode records.")
    
    category = request.session.get('create_category')
    if not category:
        return redirect('create_step1')
    
    if category == 'permit':
        types = [
            {'value': 'Building', 'label': 'Building Permit', 'icon': 'building-2', 'desc': 'Standard building permit structure approvals.'},
            {'value': 'Electrical', 'label': 'Electrical Permit', 'icon': 'zap', 'desc': 'Electrical wiring and electrical installation approvals.'},
            {'value': 'Occupancy', 'label': 'Occupancy Permit', 'icon': 'check-square', 'desc': 'Certificate of occupancy approvals.'},
            {'value': 'Fencing', 'label': 'Fencing Permit', 'icon': 'fence', 'desc': 'Fencing installation clearances.'},
            {'value': 'Demolition', 'label': 'Demolition Permit', 'icon': 'trash-2', 'desc': 'Demolition and site clearing approvals.'},
            {'value': 'Renovation', 'label': 'Renovation Permit', 'icon': 'file-text', 'desc': 'Minor renovation and building repair approvals.'},
        ]
    else:
        types = [
            {'value': 'Road & Bridge', 'label': 'Road & Bridge', 'icon': 'milestone', 'desc': 'Road concreting, bridges, and pathways.'},
            {'value': 'Building', 'label': 'Building', 'icon': 'building', 'desc': 'Government buildings, gyms, or centers.'},
            {'value': 'Water System', 'label': 'Water System', 'icon': 'droplets', 'desc': 'Water lines, wells, and irrigation projects.'},
            {'value': 'Flood Control', 'label': 'Flood Control', 'icon': 'shield-alert', 'desc': 'Seawalls, dikes, and revetments.'},
            {'value': 'Drainage', 'label': 'Drainage', 'icon': 'git-commit', 'desc': 'Drainage lines and culverts.'},
            {'value': 'Multi-purpose Hall', 'label': 'Multi-purpose Hall', 'icon': 'home', 'desc': 'Community halls and gymnasiums.'},
            {'value': 'Others', 'label': 'Others', 'icon': 'folder', 'desc': 'Other public infrastructure works.'},
        ]
    
    if request.method == 'POST':
        subtype = request.POST.get('record_subtype')
        if subtype:
            request.session['create_subtype'] = subtype
            return redirect('create_step3')
        messages.error(request, "Please select a type.")
        
    return render(request, 'permits/create_step2.html', {
        'category': category,
        'types': types,
        'active_tab': 'records'
    })

@login_required
def record_create_step3_view(request):
    if request.user.role not in ['staff', 'admin']:
        raise PermissionDenied("You do not have permission to encode records.")
    
    category = request.session.get('create_category')
    subtype = request.session.get('create_subtype')
    if not category or not subtype:
        return redirect('create_step1')
    
    barangays = Barangay.objects.all()
    scope = 'Municipal' if category == 'municipal' else ('Barangay' if category == 'barangay' else '')
    record_type = 'Permit' if category == 'permit' else 'Project'
    
    template = RequirementTemplate.objects.filter(
        record_type=record_type, subtype=subtype, scope=scope, is_active=True
    ).first()
    
    if request.method == 'POST':
        barangay_id = request.POST.get('barangay')
        year = request.POST.get('year') or None
        status = request.POST.get('status', 'active')
        
        current_year = timezone.now().year
        if year:
            try:
                year_val = int(year)
                if year_val > current_year:
                    messages.error(request, f"Filing year cannot be in the future (max {current_year}).")
                    return render(request, 'permits/create_step3.html', {
                        'category': category,
                        'subtype': subtype,
                        'scope': scope,
                        'barangays': barangays,
                        'template': template,
                        'current_year': current_year,
                        'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
                        'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
                        'status_choices': EngineeringRecord.STATUS_CHOICES,
                        'active_tab': 'records'
                    })
            except ValueError:
                messages.error(request, "Invalid year value.")
                return render(request, 'permits/create_step3.html', {
                    'category': category,
                    'subtype': subtype,
                    'scope': scope,
                    'barangays': barangays,
                    'template': template,
                    'current_year': current_year,
                    'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
                    'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
                    'status_choices': EngineeringRecord.STATUS_CHOICES,
                    'active_tab': 'records'
                })
        
        if category == 'permit':
            applicant_name = sanitize_input(request.POST.get('applicant_name', '')).strip()
            permit_number = sanitize_input(request.POST.get('permit_number', '')).strip()
            title = permit_number + ' — ' + applicant_name if permit_number else applicant_name
        else:
            title = sanitize_input(request.POST.get('title', '')).strip()
            
        if not barangay_id or not title or not year:
            messages.error(request, "Please fill in all required fields.")
            return render(request, 'permits/create_step3.html', {
                'category': category,
                'subtype': subtype,
                'scope': scope,
                'barangays': barangays,
                'template': template,
                'current_year': timezone.now().year,
                'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
                'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
                'status_choices': EngineeringRecord.STATUS_CHOICES,
                'active_tab': 'records'
            })
            
        is_illegal = request.POST.get('is_illegal_construction') == 'on' or request.POST.get('is_illegal_construction') == 'true'
        illegal_status = request.POST.get('illegal_compliance_status', 'unresolved') if is_illegal else 'unresolved'

        record = EngineeringRecord.objects.create(
            record_type=record_type,
            project_scope=scope,
            barangay_id=barangay_id,
            title=title,
            year=year,
            description=sanitize_input(request.POST.get('description', '')).strip() if category != 'permit' else '',
            status=status,
            is_illegal_construction=is_illegal,
            illegal_compliance_status=illegal_status,
            created_by=request.user,
        )
        
        if record_type == 'Permit':
            PermitDetail.objects.create(
                engineering_record=record,
                permit_type=subtype,
                building_type=request.POST.get('building_type', ''),
                permit_number=permit_number,
                applicant_name=applicant_name,
                remarks=sanitize_input(request.POST.get('remarks', '')).strip(),
            )
        elif record_type == 'Project':
            project_status = request.POST.get('project_status', 'Planning')
            ProjectDetail.objects.create(
                engineering_record=record,
                project_type=subtype,
                funding_source=sanitize_input(request.POST.get('funding_source', '')).strip(),
                contractor=sanitize_input(request.POST.get('contractor', '')).strip(),
                project_cost=request.POST.get('project_cost', '') or None,
                project_status=project_status,
            )
            # Sync parent record status
            if project_status == 'Completed':
                record.status = 'completed'
            elif project_status == 'Ongoing':
                record.status = 'in_progress'
            else:
                record.status = 'active'
            record.save()
            
        if template:
            RecordRequirement.objects.bulk_create([
                RecordRequirement(record=record, requirement_item=item)
                for item in template.active_items
            ])
            
        request.session.pop('create_category', None)
        request.session.pop('create_subtype', None)
        
        log_audit(
            request.user,
            f"Created {record_type} record: '{title}'",
            record.record_id, request
        )
        messages.success(request, f"Record '{title}' created successfully! Check list is ready.")
        return redirect('record_detail', record_id=record.record_id)
        
    return render(request, 'permits/create_step3.html', {
        'category': category,
        'subtype': subtype,
        'scope': scope,
        'barangays': barangays,
        'template': template,
        'current_year': timezone.now().year,
        'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
        'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'active_tab': 'records'
    })



# ─── MODULE LIST VIEWS ────────────────────────────────────────────────────────

@login_required
def municipal_projects_view(request):
    """Lists all Municipal Project records."""
    records = EngineeringRecord.objects.filter(
        record_type='Project', project_scope='Municipal'
    ).exclude(status='archived').select_related('barangay', 'created_by', 'project_detail').order_by('-created_at')

    query = request.GET.get('q', '').strip()
    project_type = request.GET.get('project_type', '')
    status = request.GET.get('status', '')
    year = request.GET.get('year', '')
    barangay_id = request.GET.get('barangay', '')

    if query:
        search_filter = (
            Q(title__icontains=query) |
            Q(barangay__barangay_name__icontains=query) |
            Q(project_detail__contractor__icontains=query)
        )
        if query.isdigit():
            search_filter |= Q(year=int(query))
        records = records.filter(search_filter).distinct()
    if project_type:
        records = records.filter(project_detail__project_type=project_type)
    if status:
        records = records.filter(status=status)
    if year:
        try:
            records = records.filter(year=int(year))
        except (ValueError, TypeError):
            records = records.filter(year=year)
    if barangay_id:
        records = records.filter(barangay_id=barangay_id)

    per_page = get_per_page(request, 10)
    paginator = Paginator(records, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    year_choices = get_year_choices()
    active_filters_count = sum(1 for val in [project_type, status, year, barangay_id] if val)

    context = {
        'per_page': per_page,
        'page_obj': page_obj,
        'q': query,
        'selected_project_type': project_type,
        'selected_status': status,
        'selected_year': year,
        'selected_barangay': barangay_id,
        'project_types': ProjectDetail.PROJECT_TYPE_CHOICES,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'barangays': Barangay.objects.all(),
        'year_choices': year_choices,
        'active_filters_count': active_filters_count,
        'module_title': 'Municipal Projects',
        'module_scope': 'Municipal',
        'active_tab': 'municipal',
    }
    return render(request, 'permits/module_projects.html', context)


@login_required
def barangay_projects_view(request):
    """Lists all Barangay Project records."""
    records = EngineeringRecord.objects.filter(
        record_type='Project', project_scope='Barangay'
    ).exclude(status='archived').select_related('barangay', 'created_by', 'project_detail').order_by('-created_at')

    query = request.GET.get('q', '').strip()
    project_type = request.GET.get('project_type', '')
    status = request.GET.get('status', '')
    year = request.GET.get('year', '')
    barangay_id = request.GET.get('barangay', '')

    if query:
        search_filter = (
            Q(title__icontains=query) |
            Q(barangay__barangay_name__icontains=query) |
            Q(project_detail__contractor__icontains=query)
        )
        if query.isdigit():
            search_filter |= Q(year=int(query))
        records = records.filter(search_filter).distinct()
    if project_type:
        records = records.filter(project_detail__project_type=project_type)
    if status:
        records = records.filter(status=status)
    if year:
        try:
            records = records.filter(year=int(year))
        except (ValueError, TypeError):
            records = records.filter(year=year)
    if barangay_id:
        records = records.filter(barangay_id=barangay_id)

    per_page = get_per_page(request, 10)
    paginator = Paginator(records, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    year_choices = get_year_choices()
    active_filters_count = sum(1 for val in [project_type, status, year, barangay_id] if val)

    context = {
        'per_page': per_page,
        'page_obj': page_obj,
        'q': query,
        'selected_project_type': project_type,
        'selected_status': status,
        'selected_year': year,
        'selected_barangay': barangay_id,
        'project_types': ProjectDetail.PROJECT_TYPE_CHOICES,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'barangays': Barangay.objects.all(),
        'year_choices': year_choices,
        'active_filters_count': active_filters_count,
        'module_title': 'Barangay Projects',
        'module_scope': 'Barangay',
        'active_tab': 'barangay',
    }
    return render(request, 'permits/module_projects.html', context)


@login_required
def permit_records_view(request):
    """Lists all Permit records."""
    records = EngineeringRecord.objects.filter(
        record_type='Permit'
    ).exclude(status='archived').select_related('barangay', 'created_by', 'permit_detail').order_by('-created_at')

    query = request.GET.get('q', '').strip()
    permit_type = request.GET.get('permit_type', '')
    status = request.GET.get('status', '')
    year = request.GET.get('year', '')
    barangay_id = request.GET.get('barangay', '')

    if query:
        search_filter = (
            Q(title__icontains=query) |
            Q(permit_detail__applicant_name__icontains=query) |
            Q(permit_detail__permit_number__icontains=query) |
            Q(barangay__barangay_name__icontains=query)
        )
        if query.isdigit():
            search_filter |= Q(year=int(query))
        records = records.filter(search_filter).distinct()
    if permit_type:
        records = records.filter(permit_detail__permit_type=permit_type)
    if status:
        records = records.filter(status=status)
    if year:
        try:
            records = records.filter(year=int(year))
        except (ValueError, TypeError):
            records = records.filter(year=year)
    if barangay_id:
        records = records.filter(barangay_id=barangay_id)

    # Count complete vs pending permits for status line
    pending_count = 0
    complete_count = 0
    for r in records:
        if r.completion_stats['is_complete']:
            complete_count += 1
        else:
            pending_count += 1

    per_page = get_per_page(request, 10)
    paginator = Paginator(records, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    year_choices = get_year_choices()
    active_filters_count = sum(1 for val in [permit_type, status, year, barangay_id] if val)

    context = {
        'per_page': per_page,
        'page_obj': page_obj,
        'q': query,
        'selected_permit_type': permit_type,
        'selected_status': status,
        'selected_year': year,
        'selected_barangay': barangay_id,
        'permit_types': PermitDetail.PERMIT_TYPE_CHOICES,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'barangays': Barangay.objects.all(),
        'year_choices': year_choices,
        'active_filters_count': active_filters_count,
        'module_title': 'Permit Applications',
        'active_tab': 'permits',
        'pending_count': pending_count,
        'complete_count': complete_count,
    }
    return render(request, 'permits/module_permits.html', context)


# ─── CREATE RECORD (SINGLE PAGE FORM) ──────────────────────────────────────────

@login_required
def record_create_view(request):
    """Unified single-page form to create record and load checklist instantly."""
    if request.user.role not in ['staff', 'admin']:
        raise PermissionDenied("You do not have permission to encode records.")

    barangays = Barangay.objects.all()

    if request.method == 'POST':
        category = request.POST.get('category', '')
        subtype = request.POST.get('subtype', '')
        barangay_id = request.POST.get('barangay', '')
        year = request.POST.get('year', '') or None
        status = request.POST.get('status', 'active')
        
        current_year = timezone.now().year
        if year:
            try:
                year_val = int(year)
                if year_val > current_year:
                    messages.error(request, f"Filing year cannot be in the future (max {current_year}).")
                    return render(request, 'permits/create_record.html', {
                        'barangays': barangays,
                        'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
                        'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
                        'status_choices': EngineeringRecord.STATUS_CHOICES,
                        'current_year': current_year,
                        'active_tab': 'records',
                    })
            except ValueError:
                messages.error(request, "Invalid year value.")
                return render(request, 'permits/create_record.html', {
                    'barangays': barangays,
                    'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
                    'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
                    'status_choices': EngineeringRecord.STATUS_CHOICES,
                    'current_year': current_year,
                    'active_tab': 'records',
                })

        # Resolve title
        if category == 'permit':
            applicant_name = sanitize_input(request.POST.get('applicant_name', '')).strip()
            permit_number = sanitize_input(request.POST.get('permit_number', '')).strip()
            title = sanitize_input(request.POST.get('permit_title', '')).strip()
            if not title:
                title = permit_number + ' — ' + applicant_name if permit_number else applicant_name
            record_type = 'Permit'
            scope = ''
        else:
            title = sanitize_input(request.POST.get('title', '')).strip()
            record_type = 'Project'
            scope = 'Municipal' if category == 'municipal' else 'Barangay'

        if not category or not subtype or not barangay_id or not title:
            messages.error(request, "Please fill in all required fields.")
            return render(request, 'permits/create_record.html', {
                'barangays': barangays,
                'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
                'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
                'status_choices': EngineeringRecord.STATUS_CHOICES,
                'current_year': timezone.now().year,
                'active_tab': 'records',
            })

        is_illegal = request.POST.get('is_illegal_construction') == 'on' or request.POST.get('is_illegal_construction') == 'true'
        illegal_status = request.POST.get('illegal_compliance_status', 'unresolved') if is_illegal else 'unresolved'

        # Save record
        record = EngineeringRecord.objects.create(
            record_type=record_type,
            project_scope=scope,
            barangay_id=barangay_id,
            title=title,
            year=year,
            description=sanitize_input(request.POST.get('description', '')).strip() if category != 'permit' else '',
            status=status,
            is_illegal_construction=is_illegal,
            illegal_compliance_status=illegal_status,
            created_by=request.user,
        )

        # Save details
        if record_type == 'Permit':
            PermitDetail.objects.create(
                engineering_record=record,
                permit_type=subtype,
                building_type=request.POST.get('building_type', ''),
                permit_number=permit_number,
                applicant_name=applicant_name,
                remarks=sanitize_input(request.POST.get('remarks', '')).strip(),
            )
        elif record_type == 'Project':
            project_status = request.POST.get('project_status', 'Planning')
            ProjectDetail.objects.create(
                engineering_record=record,
                project_type=subtype,
                funding_source=sanitize_input(request.POST.get('funding_source', '')).strip(),
                contractor=sanitize_input(request.POST.get('contractor', '')).strip(),
                project_cost=request.POST.get('project_cost', '') or None,
                project_status=project_status,
            )
            # Sync parent record status based on project status
            if project_status == 'Completed':
                record.status = 'completed'
            elif project_status == 'Ongoing':
                record.status = 'in_progress'
            else:
                record.status = 'active'
            record.save()

        # Generate Checklist requirements from template
        template = RequirementTemplate.objects.filter(
            record_type=record_type, subtype=subtype, scope=scope, is_active=True
        ).first()
        if template:
            RecordRequirement.objects.bulk_create([
                RecordRequirement(record=record, requirement_item=item)
                for item in template.active_items
            ])

        log_audit(
            request.user,
            f"Created {record_type} record: '{title}'",
            record.record_id, request
        )
        messages.success(request, f"Record '{title}' created successfully! Check list is ready.")
        return redirect('record_detail', record_id=record.record_id)

    return render(request, 'permits/create_record.html', {
        'barangays': barangays,
        'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
        'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'current_year': timezone.now().year,
        'active_tab': 'records',
    })




# ─── RECORD DETAIL ───────────────────────────────────────────────────────────

@login_required
def record_detail_view(request, record_id):
    record = get_object_or_404(EngineeringRecord, record_id=record_id)

    # Auto-populate checklist requirements if missing
    if not record.requirements.exists():
        template = None
        if record.record_type == 'Permit':
            subtype = record.permit_detail.permit_type if hasattr(record, 'permit_detail') and record.permit_detail and record.permit_detail.permit_type else 'Building'
            template = RequirementTemplate.objects.filter(record_type='Permit', subtype=subtype, is_active=True).first()
            if not template:
                template = RequirementTemplate.objects.filter(record_type='Permit', subtype='Building', is_active=True).first()
        elif record.record_type == 'Project':
            subtype = record.project_detail.project_type if hasattr(record, 'project_detail') and record.project_detail and record.project_detail.project_type else 'Road & Bridge'
            scope = record.project_scope or 'Municipal'
            template = RequirementTemplate.objects.filter(record_type='Project', subtype=subtype, scope=scope, is_active=True).first()

        if template:
            RecordRequirement.objects.bulk_create([
                RecordRequirement(record=record, requirement_item=item)
                for item in template.active_items
            ])

    # Load checklist requirements with their linked documents
    requirements = record.requirements.select_related(
        'requirement_item', 'document', 'fulfilled_by'
    ).order_by('requirement_item__order', 'requirement_item__name')

    # Completion stats
    completion = record.completion_stats

    # Non-checklist documents (uploaded without a slot)
    documents = record.documents.filter(requirement_item__isnull=True).order_by('-uploaded_at')

    # Generate signed URLs for all documents
    all_docs = record.documents.all()
    doc_url_map = {}
    for doc in all_docs:
        doc_url_map[doc.document_id] = reverse('serve_document', kwargs={
            'token': signing.dumps({'document_id': doc.document_id}, salt='document-download')
        })

    # Get detail
    permit_detail = None
    project_detail = None
    if record.record_type == 'Permit':
        try:
            permit_detail = record.permit_detail
        except PermitDetail.DoesNotExist:
            pass
    elif record.record_type == 'Project':
        try:
            project_detail = record.project_detail
        except ProjectDetail.DoesNotExist:
            pass

    # Activity timeline
    timeline = AuditLog.objects.filter(
        target_record_id=record_id
    ).select_related('user').order_by('-performed_at')

    # Related records (same barangay, same type)
    related_records = EngineeringRecord.objects.filter(
        barangay=record.barangay,
        record_type=record.record_type,
    ).exclude(record_id=record.record_id).select_related('barangay')[:4]

    import datetime
    today = timezone.now().date()
    thirty_days_later = today + datetime.timedelta(days=30)

    context = {
        'record': record,
        'requirements': requirements,
        'completion': completion,
        'documents': documents,
        'doc_url_map': doc_url_map,
        'permit_detail': permit_detail,
        'project_detail': project_detail,
        'timeline': timeline,
        'related_records': related_records,
        'can_edit': (request.user.role == 'admin' or (request.user.role == 'staff' and record.created_by == request.user)),
        'can_archive': (request.user.role == 'admin'),
        'active_tab': 'records',
        'today': today,
        'thirty_days_later': thirty_days_later,
    }
    return render(request, 'permits/record_detail.html', context)


@login_required
def update_illegal_status_view(request, record_id):
    """Updates illegal construction status / compliance regularization status for a record."""
    if request.user.role not in ['staff', 'admin']:
        return HttpResponseForbidden("Unauthorized")
    record = get_object_or_404(EngineeringRecord, record_id=record_id)
    if request.method == 'POST':
        status_val = request.POST.get('illegal_compliance_status', '').strip()
        flag_val = request.POST.get('is_illegal_construction', '')
        
        if flag_val == 'toggle':
            record.is_illegal_construction = not record.is_illegal_construction
            if record.is_illegal_construction and not record.illegal_compliance_status:
                record.illegal_compliance_status = 'unresolved'
            record.save()
            action_msg = "Flagged as Illegal Construction." if record.is_illegal_construction else "Unflagged Illegal Construction."
            log_audit(request.user, action_msg, target_record_id=record.record_id, request=request)
            messages.success(request, action_msg)
        elif status_val in ['unresolved', 'pending_permit', 'resolved']:
            record.is_illegal_construction = True
            record.illegal_compliance_status = status_val
            record.save()
            lbl = record.get_illegal_compliance_status_display()
            log_audit(request.user, f"Updated Illegal Construction Compliance to {lbl}", target_record_id=record.record_id, request=request)
            messages.success(request, f"Regularization status updated to '{lbl}'.")
        elif status_val == 'remove':
            record.is_illegal_construction = False
            record.save()
            log_audit(request.user, "Removed Illegal Construction flag", target_record_id=record.record_id, request=request)
            messages.success(request, "Illegal construction flag removed.")
            
    return redirect('record_detail', record_id=record.record_id)


@login_required
def flag_illegal_construction_view(request):
    """Handles direct reporting/flagging of an unpermitted illegal construction structure."""
    if request.user.role not in ['staff', 'admin']:
        return HttpResponseForbidden("Unauthorized")
    
    if request.method == 'POST':
        title = sanitize_input(request.POST.get('title', '')).strip()
        barangay_id = request.POST.get('barangay', '')
        location_address = sanitize_input(request.POST.get('location_address', '')).strip()
        date_discovered_str = request.POST.get('date_discovered', '')
        status_val = request.POST.get('illegal_compliance_status', 'unresolved')
        
        if not title:
            title = "Unpermitted Structure Discovered"
        if not barangay_id:
            messages.error(request, "Please select a barangay.")
            return redirect(request.META.get('HTTP_REFERER', 'records_browse'))
            
        barangay = get_object_or_404(Barangay, barangay_id=barangay_id)
        
        # Parse date discovered or default to today
        if date_discovered_str:
            try:
                from datetime import datetime
                date_discovered = datetime.strptime(date_discovered_str, '%Y-%m-%d').date()
            except ValueError:
                date_discovered = timezone.now().date()
        else:
            date_discovered = timezone.now().date()
            
        record = EngineeringRecord.objects.create(
            record_type='Permit',
            project_scope='',
            barangay=barangay,
            title=title,
            year=date_discovered.year,
            description=location_address,
            status='active',
            date_started=date_discovered,
            is_illegal_construction=True,
            illegal_compliance_status=status_val if status_val in ['unresolved', 'pending_permit', 'resolved'] else 'unresolved',
            created_by=request.user
        )
        
        PermitDetail.objects.create(
            engineering_record=record,
            permit_type='Building',
            building_type='Commercial',
            permit_number='',
            applicant_name='[Unpermitted Construction Discovered]',
            remarks=f"Flagged as unpermitted structure on {date_discovered.strftime('%d %b %Y')}. Location: {location_address}"
        )
        
        # Attach Building Permit / Regularization Checklist Template
        template = RequirementTemplate.objects.filter(record_type='Permit', subtype='Building', is_active=True).first()
        if template:
            RecordRequirement.objects.bulk_create([
                RecordRequirement(record=record, requirement_item=item)
                for item in template.active_items
            ])

        # Handle Discovery Photo upload
        if 'photo' in request.FILES and request.FILES['photo']:
            photo_file = request.FILES['photo']
            try:
                from django.core.exceptions import ValidationError
                validate_document_file(photo_file)
                Document.objects.create(
                    engineering_record=record,
                    document_type='Picture',
                    file_name=photo_file.name,
                    file=photo_file,
                    file_size=photo_file.size,
                    uploaded_by=request.user
                )
            except ValidationError as err:
                messages.warning(request, f"Flagged illegal construction, but photo attachment failed: {err.message if hasattr(err, 'message') else str(err)}")
                
        status_lbl = record.get_illegal_compliance_status_display()
        log_audit(
            request.user,
            f"Flagged Illegal Construction at Barangay {barangay.barangay_name}: '{title}' (Status: {status_lbl})",
            target_record_id=record.record_id,
            request=request
        )
        messages.success(request, f"Successfully reported/flagged unpermitted structure in Barangay {barangay.barangay_name}.")
        return redirect('record_detail', record_id=record.record_id)

    return redirect('records_browse')



@login_required
def toggle_requirement_waived_view(request, req_id):
    """Toggles the waived (N/A) status of a specific record requirement."""
    if request.user.role not in ['staff', 'admin']:
        return HttpResponseForbidden("Unauthorized")
        
    req = get_object_or_404(RecordRequirement, req_id=req_id)
    
    # Permission check: staff can only toggle requirements for records they created
    if request.user.role == 'staff' and req.record.created_by != request.user:
        return HttpResponseForbidden("You can only edit requirements for records you created.")
    
    req.is_waived = not req.is_waived
    req.save()
    
    action_str = "marked as N/A" if req.is_waived else "marked as required"
    messages.success(request, f"Requirement '{req.requirement_item.name}' successfully {action_str}.")
    log_audit(
        request.user, 
        f"Requirement '{req.requirement_item.name}' {action_str} for record '{req.record.title}'", 
        target_record_id=req.record.record_id, 
        request=request
    )
    
    stats = req.record.completion_stats
    return JsonResponse({
        'success': True,
        'is_waived': req.is_waived,
        'fulfilled': stats['fulfilled'],
        'total': stats['total'],
        'pct': stats['pct'],
        'is_complete': stats['is_complete']
    })


# ─── RECORD EDIT ─────────────────────────────────────────────────────────────

@login_required
def record_edit_view(request, record_id):
    record = get_object_or_404(EngineeringRecord, record_id=record_id)

    # Permission check
    if request.user.role == 'staff' and record.created_by != request.user:
        raise PermissionDenied("You can only edit your own records.")
    if request.user.role not in ['staff', 'admin']:
        raise PermissionDenied("You do not have permission to edit records.")

    barangays = Barangay.objects.all()

    if request.method == 'POST':
        record.title = sanitize_input(request.POST.get('title', '')).strip()
        record.description = sanitize_input(request.POST.get('description', '')).strip()
        record.status = request.POST.get('status', record.status)
        record.barangay_id = request.POST.get('barangay', record.barangay_id)
        record.date_started = request.POST.get('date_started', '') or None
        record.date_completed = request.POST.get('date_completed', '') or None
        
        is_illegal = request.POST.get('is_illegal_construction') == 'on' or request.POST.get('is_illegal_construction') == 'true'
        record.is_illegal_construction = is_illegal
        if is_illegal:
            record.illegal_compliance_status = request.POST.get('illegal_compliance_status', record.illegal_compliance_status or 'unresolved')
        else:
            record.illegal_compliance_status = 'unresolved'
            
        record.save()

        # Update detail and regenerate requirements if subtype changed or doesn't exist
        if record.record_type == 'Permit':
            detail, _ = PermitDetail.objects.get_or_create(engineering_record=record)
            old_subtype = detail.permit_type
            new_subtype = request.POST.get('permit_type', '')
            
            detail.permit_type = new_subtype
            detail.building_type = request.POST.get('building_type', detail.building_type)
            detail.permit_number = sanitize_input(request.POST.get('permit_number', '')).strip()
            detail.applicant_name = sanitize_input(request.POST.get('applicant_name', '')).strip()
            detail.resolution_required = request.POST.get('resolution_required') == 'on'
            detail.remarks = sanitize_input(request.POST.get('remarks', '')).strip()
            detail.save()
            
            if not record.requirements.exists() or old_subtype != new_subtype:
                record.requirements.all().delete()
                template = RequirementTemplate.objects.filter(
                    record_type='Permit', subtype=new_subtype, is_active=True
                ).first()
                if template:
                    RecordRequirement.objects.bulk_create([
                        RecordRequirement(record=record, requirement_item=item)
                        for item in template.active_items
                    ])
                    
        elif record.record_type == 'Project':
            detail, _ = ProjectDetail.objects.get_or_create(engineering_record=record)
            old_subtype = detail.project_type
            new_subtype = request.POST.get('project_type', '')
            
            detail.project_type = new_subtype
            detail.funding_source = sanitize_input(request.POST.get('funding_source', '')).strip()
            detail.contractor = sanitize_input(request.POST.get('contractor', '')).strip()
            detail.project_cost = request.POST.get('project_cost', '') or None
            project_status = request.POST.get('project_status', detail.project_status)
            detail.project_status = project_status
            detail.save()
            
            # Sync parent record status based on project status
            if project_status == 'Completed':
                record.status = 'completed'
            elif project_status == 'Ongoing':
                record.status = 'in_progress'
            else:
                if record.status not in ['archived', 'pending']:
                    record.status = 'active'
            record.save()
            
            if not record.requirements.exists() or old_subtype != new_subtype:
                record.requirements.all().delete()
                template = RequirementTemplate.objects.filter(
                    record_type='Project', subtype=new_subtype, scope=record.project_scope, is_active=True
                ).first()
                if template:
                    RecordRequirement.objects.bulk_create([
                        RecordRequirement(record=record, requirement_item=item)
                        for item in template.active_items
                    ])

        log_audit(request.user, f"Updated {record.record_type}: '{record.title}'", record.record_id, request)
        messages.success(request, f"Record '{record.title}' updated successfully.")
        return redirect('record_detail', record_id=record.record_id)

    # Get detail for form
    permit_detail = None
    project_detail = None
    if record.record_type == 'Permit':
        try:
            permit_detail = record.permit_detail
        except PermitDetail.DoesNotExist:
            pass
    elif record.record_type == 'Project':
        try:
            project_detail = record.project_detail
        except ProjectDetail.DoesNotExist:
            pass

    context = {
        'record': record,
        'permit_detail': permit_detail,
        'project_detail': project_detail,
        'barangays': barangays,
        'permit_types': PermitDetail.PERMIT_TYPE_CHOICES,
        'building_types': PermitDetail.BUILDING_TYPE_CHOICES,
        'project_types': ProjectDetail.PROJECT_TYPE_CHOICES,
        'project_statuses': ProjectDetail.PROJECT_STATUS_CHOICES,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'active_tab': 'records',
    }
    return render(request, 'permits/edit_record.html', context)


# ─── SERVE SECURE DOCUMENT ──────────────────────────────────────────────────

@login_required
def serve_document_view(request, token):
    doc = None
    if str(token).isdigit():
        doc = get_object_or_404(Document, document_id=int(token))
    else:
        try:
            data = signing.loads(token, salt='document-download', max_age=600)
            doc = get_object_or_404(Document, document_id=data['document_id'])
        except (signing.SignatureExpired, signing.BadSignature):
            try:
                doc = get_object_or_404(Document, document_id=token)
            except Exception:
                return HttpResponseForbidden("Invalid document token link.")

    if not doc or not doc.file:
        raise Http404("Document file not found.")

    try:
        if hasattr(doc.file, 'url') and str(doc.file.url).startswith('http'):
            return redirect(doc.file.url)
        response = FileResponse(doc.file.open('rb'), content_type='application/octet-stream')
        response['Content-Disposition'] = f'inline; filename="{doc.file_name}"'
        return response
    except Exception as exc:
        logger.error(f"Error serving document file: {exc}")
        if hasattr(doc.file, 'url') and doc.file.url:
            return redirect(doc.file.url)
        raise Http404("Unable to access stored document file.")


# ─── DOCUMENT UPLOAD / DELETE ────────────────────────────────────────────────

@login_required
def document_upload_view(request, record_id):
    record = get_object_or_404(EngineeringRecord, record_id=record_id)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'json' in request.headers.get('Accept', '').lower()

    if request.user.role not in ['staff', 'admin']:
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'You do not have permission to upload documents.'}, status=403)
        raise PermissionDenied("You do not have permission to upload documents.")

    if request.method == 'POST':
        document_file = request.FILES.get('document_file')
        requirement_item_id = request.POST.get('requirement_item_id', '').strip()
        document_type = request.POST.get('document_type', 'Other').strip()
        expiry_date = request.POST.get('expiry_date', '').strip()

        if not document_file:
            err_msg = "Please select a file to upload."
            if is_ajax:
                return JsonResponse({'success': False, 'error': err_msg}, status=400)
            messages.error(request, err_msg)
            return redirect('record_detail', record_id=record.record_id)

        try:
            validate_document_file(document_file)
        except Exception as exc:
            err_msg = exc.message if hasattr(exc, 'message') else str(exc)
            if is_ajax:
                return JsonResponse({'success': False, 'error': err_msg}, status=400)
            messages.error(request, err_msg)
            return redirect('record_detail', record_id=record.record_id)

        import datetime
        parsed_expiry_date = None
        if expiry_date:
            try:
                parsed_expiry_date = datetime.datetime.strptime(expiry_date, '%Y-%m-%d').date()
            except ValueError:
                pass

        req_item = None
        if requirement_item_id:
            try:
                req_item = RequirementItem.objects.get(item_id=requirement_item_id)
                document_type = req_item.name  # use the requirement name as document type
                # Prevent orphan file/record leaks by checking for existing documents in this slot
                existing_req = RecordRequirement.objects.filter(record=record, requirement_item=req_item).first()
                if existing_req and existing_req.document:
                    old_doc = existing_req.document
                    try:
                        old_doc.file.delete()
                        old_doc.delete()
                    except Exception as e:
                        logger.error(f"Error deleting replaced document: {e}")
            except RequirementItem.DoesNotExist:
                pass

        try:
            doc = Document.objects.create(
                engineering_record=record,
                requirement_item=req_item,
                document_type=document_type,
                file=document_file,
                file_name=document_file.name,
                file_size=document_file.size,
                uploaded_by=request.user,
                expiry_date=parsed_expiry_date,
            )
        except Exception as exc:
            logger.error(f"Failed saving document file to storage: {exc}")
            err_msg = f"Storage Save Error: {str(exc)}"
            if is_ajax:
                return JsonResponse({'success': False, 'error': err_msg}, status=500)
            messages.error(request, err_msg)
            return redirect('record_detail', record_id=record.record_id)

        # Mark the corresponding checklist slot as fulfilled
        if req_item:
            RecordRequirement.objects.filter(
                record=record, requirement_item=req_item
            ).update(
                document=doc,
                is_fulfilled=True,
                fulfilled_at=timezone.now(),
                fulfilled_by=request.user,
            )

        log_audit(
            request.user,
            f"Uploaded: {req_item.name if req_item else 'general'} for '{record.title}'",
            record.record_id, request
        )
        msg_str = f"Uploaded: {req_item.name if req_item else 'general'} successfully."
        messages.success(request, msg_str)

        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': msg_str,
                'doc_id': doc.document_id,
                'file_name': doc.file_name,
                'file_url': f"/documents/serve/{doc.document_id}/"
            })

    return redirect('record_detail', record_id=record.record_id)


@login_required
def document_delete_view(request, record_id, document_id):
    record = get_object_or_404(EngineeringRecord, record_id=record_id)
    doc = get_object_or_404(Document, document_id=document_id, engineering_record=record)

    if request.user.role != 'admin':
        raise PermissionDenied("Only Administrators can delete documents.")

    # Mark the corresponding checklist requirement slot as unfulfilled
    RecordRequirement.objects.filter(document=doc).update(
        document=None,
        is_fulfilled=False,
        fulfilled_at=None,
        fulfilled_by=None
    )

    doc.file.delete()
    doc.delete()
    doc_label = doc.requirement_item.name if doc.requirement_item else doc.document_type
    log_audit(request.user, f"Deleted: {doc_label} from '{record.title}'", record.record_id, request)
    messages.success(request, f"Deleted: {doc_label} successfully.")
    return redirect('record_detail', record_id=record.record_id)


# ─── ARCHIVE / RESTORE ──────────────────────────────────────────────────────

@login_required
def record_archive_view(request, record_id):
    record = get_object_or_404(EngineeringRecord, record_id=record_id)
    if request.user.role != 'admin':
        raise PermissionDenied("Only Administrators can archive records.")

    record.status = 'archived'
    record.save()
    log_audit(request.user, f"Archived: '{record.title}'", record.record_id, request)
    messages.success(request, f"Record '{record.title}' archived.")
    return redirect('records_browse')


@login_required
def record_restore_view(request, record_id):
    record = get_object_or_404(EngineeringRecord, record_id=record_id)
    if request.user.role != 'admin':
        raise PermissionDenied("Only Administrators can restore records.")

    record.status = 'active'
    record.save()
    log_audit(request.user, f"Restored: '{record.title}'", record.record_id, request)
    messages.success(request, f"Record '{record.title}' restored.")
    return redirect('record_detail', record_id=record.record_id)


# ─── ARCHIVE PAGE ───────────────────────────────────────────────────────────

@login_required
def archive_view(request):
    if request.user.role != 'admin':
        raise PermissionDenied("Only Administrators can view archived records.")
    records = EngineeringRecord.objects.filter(status='archived').select_related('barangay', 'created_by')

    query = request.GET.get('q', '').strip()
    if query:
        records = records.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )

    per_page = get_per_page(request, 10)
    paginator = Paginator(records, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'per_page': per_page,
        'page_obj': page_obj,
        'q': query,
        'active_tab': 'archive',
    }
    return render(request, 'permits/archive.html', context)


# ─── SEARCH ──────────────────────────────────────────────────────────────────

@login_required
def search_view(request):
    query = request.GET.get('q', '').strip()
    record_type = request.GET.get('record_type', '').strip()
    barangay_id = request.GET.get('barangay', '').strip()
    year = request.GET.get('year', '').strip()
    status = request.GET.get('status', '').strip()

    records = EngineeringRecord.objects.exclude(status='archived').select_related(
        'barangay', 'created_by', 'permit_detail', 'project_detail'
    ).prefetch_related('documents')

    if query:
        search_filter = (
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(barangay__barangay_name__icontains=query) |
            Q(permit_detail__permit_number__icontains=query) |
            Q(permit_detail__applicant_name__icontains=query) |
            Q(permit_detail__permit_type__icontains=query) |
            Q(permit_detail__building_type__icontains=query) |
            Q(project_detail__project_type__icontains=query) |
            Q(project_detail__contractor__icontains=query) |
            Q(project_detail__funding_source__icontains=query) |
            Q(record_type__icontains=query) |
            Q(project_scope__icontains=query) |
            Q(illegal_compliance_status__icontains=query) |
            Q(status__icontains=query) |
            Q(created_by__full_name__icontains=query) |
            Q(created_by__username__icontains=query)
        )
        if query.isdigit():
            search_filter |= Q(year=int(query)) | Q(created_at__year=int(query)) | Q(date_started__year=int(query))
        records = records.filter(search_filter).distinct()

    if record_type:
        records = records.filter(record_type=record_type)
    if barangay_id:
        records = records.filter(barangay_id=barangay_id)
    if year:
        try:
            records = records.filter(year=int(year))
        except (ValueError, TypeError):
            records = records.filter(year=year)
    if status:
        records = records.filter(status=status)

    records = records.order_by('-created_at')

    per_page = get_per_page(request, 10)
    paginator = Paginator(records, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    barangays = Barangay.objects.all().order_by('barangay_name')

    context = {
        'page_obj': page_obj,
        'q': query,
        'selected_record_type': record_type,
        'selected_barangay': barangay_id,
        'selected_year': year,
        'selected_status': status,
        'barangays': barangays,
        'status_choices': EngineeringRecord.STATUS_CHOICES,
        'year_choices': get_year_choices(),
        'active_tab': 'search',
        'per_page': per_page,
    }
    return render(request, 'permits/search.html', context)




# ─── REPORTS ─────────────────────────────────────────────────────────────────

@login_required
def reports_view(request):
    """Generates detailed statistics and groupings for engineering records."""
    if request.user.role not in ['admin', 'engineer']:
        raise PermissionDenied("Only Administrators and Municipal Engineers can view summary reports.")

    # 1. Get filter parameters
    selected_record_type = request.GET.get('record_type', '').strip()
    selected_barangay = request.GET.get('barangay', '').strip()
    selected_year = request.GET.get('year', '').strip()

    # 2. Start with all records
    records = EngineeringRecord.objects.all().select_related('barangay')

    # Apply filters to base queryset
    if selected_record_type:
        records = records.filter(record_type=selected_record_type)
    if selected_barangay:
        records = records.filter(barangay_id=selected_barangay)
    if selected_year:
        records = records.filter(year=selected_year)

    # Intercept for exports
    export_format = request.GET.get('export', '').strip().lower()
    if export_format in ['excel', 'pdf']:
        # Format filters description
        meta_info = []
        if selected_record_type:
            meta_info.append(f"Type: {selected_record_type}s")
        if selected_barangay:
            barangay_obj = Barangay.objects.filter(barangay_id=selected_barangay).first()
            if barangay_obj:
                meta_info.append(f"Barangay: {barangay_obj.barangay_name}")
        if selected_year:
            meta_info.append(f"Year: {selected_year}")


        if export_format == 'excel':
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from django.http import HttpResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Engineering Records"
            ws.views.sheetView[0].showGridLines = True

            # Styles
            title_font = Font(name='Arial', size=16, bold=True, color='1E3A8A')
            header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            data_font = Font(name='Arial', size=10)
            bold_font = Font(name='Arial', size=10, bold=True)
            
            header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            subtotal_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            double_bottom_border = Border(
                top=Side(style='thin', color='94A3B8'),
                bottom=Side(style='double', color='1E3A8A')
            )

            ws['A1'] = "MUNICIPAL ENGINEERING OFFICE"
            ws['A1'].font = title_font
            ws['A2'] = f"Engineering Records Summary Report — Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            ws['A2'].font = Font(name='Arial', size=11, italic=True, color='475569')
            ws['A3'] = f"Filters: {', '.join(meta_info) if meta_info else 'All Records'}"
            ws['A3'].font = Font(name='Arial', size=9, bold=True, color='475569')

            ws.append([]) # Empty row

            headers = ["Record ID", "Reference No / Project Title", "Specific Type", "Barangay", "Year", "Status", "Owner/Client", "Estimated Cost / Fee"]
            ws.append(headers)
            
            header_row_idx = 5
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row_idx, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center' if col_idx in [1, 3, 5, 6] else 'left', vertical='center')

            total_val = 0
            for r in records:
                specific_type = r.specific_type_label
                status_label = dict(EngineeringRecord.STATUS_CHOICES).get(r.status, r.status)
                
                cost = 0
                if r.record_type == 'Project' and hasattr(r, 'project_detail') and r.project_detail:
                    cost = r.project_detail.project_cost or 0
                total_val += cost

                ref_no = r.permit_detail.permit_number if r.record_type == 'Permit' and hasattr(r, 'permit_detail') and r.permit_detail else r.title
                owner = r.permit_detail.applicant_name if r.record_type == 'Permit' and hasattr(r, 'permit_detail') and r.permit_detail else (r.project_detail.contractor if r.record_type == 'Project' and hasattr(r, 'project_detail') and r.project_detail else "")

                row_data = [r.record_id, ref_no, specific_type, r.barangay.barangay_name if r.barangay else "", r.year, status_label, owner, cost]
                ws.append(row_data)
                
                curr_row = ws.max_row
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in [1, 5, 6]:
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 8:
                        cell.alignment = Alignment(horizontal='right')
                        cell.number_format = '₱#,##0.00'

            tot_row_idx = ws.max_row + 1
            ws.cell(row=tot_row_idx, column=1, value="Total Records").font = bold_font
            ws.cell(row=tot_row_idx, column=2, value=records.count()).font = bold_font
            ws.cell(row=tot_row_idx, column=7, value="Total Cost / Fees").font = bold_font
            ws.cell(row=tot_row_idx, column=8, value=total_val).font = bold_font
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=tot_row_idx, column=col_idx)
                cell.border = double_bottom_border
                cell.fill = subtotal_fill
                if col_idx == 8:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '₱#,##0.00'

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col[4:]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=engineering_records_report.xlsx'
            wb.save(response)
            return response

        elif export_format == 'pdf':
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from io import BytesIO
            from django.http import HttpResponse

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(letter),
                rightMargin=36,
                leftMargin=36,
                topMargin=36,
                bottomMargin=36
            )
            
            story = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=16,
                textColor=colors.HexColor('#1E3A8A'),
                spaceAfter=4
            )
            meta_style = ParagraphStyle(
                'MetaStyle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=9,
                textColor=colors.HexColor('#475569'),
                spaceAfter=12
            )
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=8,
                textColor=colors.HexColor('#1E293B')
            )
            cell_center = ParagraphStyle(
                'CellCenter',
                parent=cell_style,
                alignment=1
            )
            cell_right = ParagraphStyle(
                'CellRight',
                parent=cell_style,
                alignment=2
            )
            header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=8,
                textColor=colors.white,
                alignment=1
            )

            story.append(Paragraph("MUNICIPAL ENGINEERING OFFICE", title_style))
            story.append(Paragraph(f"Engineering Records Summary Report — Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Italic']))
            meta_text = f"Filters: {', '.join(meta_info) if meta_info else 'All Records'}"
            story.append(Paragraph(meta_text, meta_style))
            story.append(Spacer(1, 10))

            col_widths = [50, 150, 100, 90, 40, 60, 110, 80]
            table_data = [[
                Paragraph("Record ID", header_style),
                Paragraph("Reference No / Project Title", header_style),
                Paragraph("Specific Type", header_style),
                Paragraph("Barangay", header_style),
                Paragraph("Year", header_style),
                Paragraph("Status", header_style),
                Paragraph("Owner/Client", header_style),
                Paragraph("Cost / Fee", header_style)
            ]]

            total_val = 0
            for r in records:
                specific_type = r.specific_type_label
                status_label = dict(EngineeringRecord.STATUS_CHOICES).get(r.status, r.status)
                
                cost = 0
                if r.record_type == 'Project' and hasattr(r, 'project_detail') and r.project_detail:
                    cost = r.project_detail.project_cost or 0
                total_val += cost

                ref_no = r.permit_detail.permit_number if r.record_type == 'Permit' and hasattr(r, 'permit_detail') and r.permit_detail else r.title
                owner = r.permit_detail.applicant_name if r.record_type == 'Permit' and hasattr(r, 'permit_detail') and r.permit_detail else (r.project_detail.contractor if r.record_type == 'Project' and hasattr(r, 'project_detail') and r.project_detail else "")

                table_data.append([
                    Paragraph(str(r.record_id), cell_center),
                    Paragraph(ref_no, cell_style),
                    Paragraph(specific_type, cell_style),
                    Paragraph(r.barangay.barangay_name if r.barangay else "", cell_style),
                    Paragraph(str(r.year), cell_center),
                    Paragraph(status_label, cell_center),
                    Paragraph(owner, cell_style),
                    Paragraph(f"Php {cost:,.2f}", cell_right)
                ])

            table_data.append([
                Paragraph("<b>Total Records:</b>", cell_style),
                Paragraph(f"<b>{records.count()}</b>", cell_style),
                Paragraph("", cell_style),
                Paragraph("", cell_style),
                Paragraph("", cell_style),
                Paragraph("", cell_style),
                Paragraph("<b>Total Cost / Fees:</b>", cell_right),
                Paragraph(f"<b>Php {total_val:,.2f}</b>", cell_right)
            ])

            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CBD5E1')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#94A3B8')),
                ('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#1E3A8A')),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ])
            t.setStyle(t_style)
            story.append(t)

            doc.build(story)
            pdf_data = buffer.getvalue()
            buffer.close()

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=engineering_records_report.pdf'
            response.write(pdf_data)
            return response

    total_count = records.count()
    total_active = records.exclude(status='archived').count()
    total_archived = records.filter(status='archived').count()

    # Category breakdown (sums exactly to total_count since all records are either Permit or Project)
    by_category = [
        {
            'category_name': 'Permit Records',
            'count': records.filter(record_type='Permit').count()
        },
        {
            'category_name': 'Municipal Projects',
            'count': records.filter(record_type='Project', project_scope='Municipal').count()
        },
        {
            'category_name': 'Barangay Projects',
            'count': records.filter(record_type='Project', project_scope='Barangay').count()
        }
    ]

    # Group by Barangay
    by_barangay = records.values('barangay__barangay_name').annotate(count=Count('record_id')).order_by('-count')

    # Group by Year
    by_year = records.values('year').annotate(count=Count('record_id')).order_by('-year')

    # Growth data (last 6 months) respecting filters
    import datetime
    growth_data = []
    today = timezone.now()
    
    # We apply category, barangay, and year filters to growth history
    growth_records = EngineeringRecord.objects.all()
    if selected_record_type:
        growth_records = growth_records.filter(record_type=selected_record_type)
    if selected_barangay:
        growth_records = growth_records.filter(barangay_id=selected_barangay)
    if selected_year:
        growth_records = growth_records.filter(year=selected_year)

    for i in range(5, -1, -1):
        year_val = today.year
        month_val = today.month - i
        if month_val <= 0:
            month_val += 12
            year_val -= 1
        first_day_of_next_month = (datetime.datetime(year_val, month_val, 1) + datetime.timedelta(days=32)).replace(day=1)
        first_day_of_next_month = timezone.make_aware(first_day_of_next_month)
        cumulative_count = growth_records.filter(created_at__lt=first_day_of_next_month).count()
        month_name = datetime.date(year_val, month_val, 1).strftime('%b %Y')
        growth_data.append({'month': month_name, 'count': cumulative_count})

    # Prepare readable filter names for UI display
    active_filters = []
    if selected_record_type:
        active_filters.append(f"Category: {selected_record_type}s")
    if selected_barangay:
        barangay_obj = Barangay.objects.filter(barangay_id=selected_barangay).first()
        if barangay_obj:
            active_filters.append(f"Barangay: {barangay_obj.barangay_name}")
    if selected_year:
        active_filters.append(f"Year: {selected_year}")

    barangays = Barangay.objects.all()

    context = {
        'total_count': total_count,
        'total_active': total_active,
        'total_archived': total_archived,
        'by_category': by_category,
        'by_barangay': by_barangay,
        'by_year': by_year,
        'growth_data': growth_data,
        'barangays': barangays,
        'selected_record_type': selected_record_type,
        'selected_barangay': selected_barangay,
        'selected_year': selected_year,
        'active_filters': active_filters,
        'active_tab': 'reports',
    }
    return render(request, 'permits/reports.html', context)



# ─── ACTIVITY LOGS ──────────────────────────────────────────────────────────

@login_required
def activity_logs_view(request):
    if request.user.role not in ['admin', 'staff', 'engineer']:
        raise PermissionDenied("You do not have permission to view activity logs.")

    if request.method == 'POST':
        if request.user.role != 'admin':
            return HttpResponseForbidden("Unauthorized action.")
        action = request.POST.get('action')
        if action == 'block_ip':
            ip = request.POST.get('ip_address')
            if ip:
                BlockedIP.objects.get_or_create(ip_address=ip, blocked_by=request.user)
                log_audit(request.user, f"Blocked IP address: {ip}", request=request)
                messages.success(request, f"Successfully blocked IP address: {ip}")
            return redirect('activity_logs')
        elif action == 'unblock_ip':
            ip = request.POST.get('ip_address')
            if ip:
                BlockedIP.objects.filter(ip_address=ip).delete()
                log_audit(request.user, f"Unblocked IP address: {ip}", request=request)
                messages.success(request, f"Successfully unblocked IP address: {ip}")
            return redirect('activity_logs')

    # 1. Audit Logs (System Activity Log)
    audit_logs = AuditLog.objects.select_related('user').order_by('-performed_at')
    if request.user.role != 'admin':
        audit_logs = audit_logs.filter(user=request.user)

    query = request.GET.get('q', '').strip()
    if query:
        audit_logs = audit_logs.filter(
            Q(action__icontains=query) | Q(user__username__icontains=query) | Q(user__email__icontains=query)
        )

    per_page = get_per_page(request, 15)
    audit_paginator = Paginator(audit_logs, per_page)
    log_page_obj = audit_paginator.get_page(request.GET.get('log_page'))

    # 2. Login History Attempts
    login_page_obj = None
    status_filter = 'all'
    if request.user.role == 'admin':
        login_attempts = LoginAttempt.objects.all().order_by('-timestamp')
        if query:
            login_attempts = login_attempts.filter(
                Q(email_attempted__icontains=query) | Q(ip_address__icontains=query)
            )

        status_filter = request.GET.get('status', 'all').strip()
        if status_filter == 'success':
            login_attempts = login_attempts.filter(success=True)
        elif status_filter == 'failed':
            login_attempts = login_attempts.filter(success=False)

        login_paginator = Paginator(login_attempts, per_page)
        login_page_obj = login_paginator.get_page(request.GET.get('login_page'))

    # Fetch currently blocked IP addresses
    blocked_ips = []
    if request.user.role == 'admin':
        blocked_ips = list(BlockedIP.objects.values_list('ip_address', flat=True))

    context = {
        'per_page': per_page,
        'log_page_obj': log_page_obj,
        'login_page_obj': login_page_obj,
        'query': query,
        'status_filter': status_filter,
        'blocked_ips': blocked_ips,
        'active_tab': 'activity_logs',
    }
    return render(request, 'permits/activity_logs.html', context)


import csv
from django.http import StreamingHttpResponse

class Echo:
    def write(self, value):
        return value

@login_required
def export_activity_logs_view(request):
    if request.user.role not in ['admin', 'staff', 'engineer']:
        raise PermissionDenied("You do not have permission to export activity logs.")
        
    tab = request.GET.get('tab', 'audit').strip()
    query = request.GET.get('q', '').strip()
    
    if tab == 'login':
        if request.user.role != 'admin':
            raise PermissionDenied("Only Administrators can export login history.")
        attempts = LoginAttempt.objects.all().order_by('-timestamp')
        if query:
            attempts = attempts.filter(
                Q(email_attempted__icontains=query) | Q(ip_address__icontains=query)
            )
            
        def login_rows():
            yield ['Timestamp', 'Email Attempted', 'IP Address', 'Result', 'Failure Reason']
            for att in attempts.iterator():
                yield [
                    att.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    att.email_attempted,
                    att.ip_address,
                    'SUCCESS' if att.success else 'FAILED',
                    att.failure_reason or ''
                ]
        
        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer)
        response = StreamingHttpResponse(
            (writer.writerow(row) for row in login_rows()),
            content_type="text/csv"
        )
        response['Content-Disposition'] = 'attachment; filename="login_history_export.csv"'
        return response
        
    else:
        logs = AuditLog.objects.select_related('user').order_by('-performed_at')
        if request.user.role != 'admin':
            logs = logs.filter(user=request.user)
        if query:
            logs = logs.filter(
                Q(action__icontains=query) | Q(user__username__icontains=query) | Q(user__email__icontains=query)
            )
            
        def audit_rows():
            yield ['Timestamp', 'User', 'Role', 'Action Executed', 'IP Address']
            for log in logs.iterator():
                yield [
                    log.performed_at.strftime('%Y-%m-%d %H:%M:%S'),
                    log.user.full_name or log.user.username,
                    log.user.get_role_display(),
                    log.action,
                    log.ip_address or ''
                ]
                
        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer)
        response = StreamingHttpResponse(
            (writer.writerow(row) for row in audit_rows()),
            content_type="text/csv"
        )
        response['Content-Disposition'] = 'attachment; filename="activity_logs_export.csv"'
        return response


# ─── PROFILE ─────────────────────────────────────────────────────────────────

@login_required
def profile_view(request):
    user = request.user
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_profile':
            full_name = sanitize_input(request.POST.get('full_name', '')).strip()
            email = sanitize_input(request.POST.get('email', '')).lower()

            if not full_name or not email:
                messages.error(request, "Full Name and Email are required.")
                return redirect('profile')

            if CustomUser.objects.exclude(id=user.id).filter(email=email).exists():
                messages.error(request, "Email already in use.")
                return redirect('profile')

            user.full_name = full_name
            user.email = email
            user.save()
            log_audit(user, "Updated profile details", request=request)
            messages.success(request, "Profile updated successfully.")
            return redirect('profile')

        elif action == 'change_password':
            current_password = request.POST.get('current_password', '')
            new_password = request.POST.get('new_password', '')
            confirm_new_password = request.POST.get('confirm_new_password', '')

            if not user.check_password(current_password):
                messages.error(request, "Current password is incorrect.")
                return redirect('profile')

            if new_password != confirm_new_password:
                messages.error(request, "New passwords do not match.")
                return redirect('profile')

            ok, err_msg = validate_password_strength(new_password)
            if not ok:
                messages.error(request, err_msg)
                return redirect('profile')

            matched_history = False
            histories = PasswordHistory.objects.filter(user=user).order_by('-created_at')[:3]
            for h in histories:
                if check_password(new_password, h.password_hash):
                    matched_history = True
                    break

            if check_password(new_password, user.password):
                matched_history = True

            if matched_history:
                messages.error(request, "Cannot reuse the last 3 passwords.")
                return redirect('profile')

            PasswordHistory.objects.create(user=user, password_hash=user.password)

            user.set_password(new_password)
            user.save()
            update_session_auth_hash(request, user)

            all_histories = PasswordHistory.objects.filter(user=user).order_by('-created_at')
            if all_histories.count() > 3:
                for h in all_histories[3:]:
                    h.delete()

            log_audit(user, "Changed password", request=request)
            messages.success(request, "Password changed successfully.")
            return redirect('profile')

    user_logs = AuditLog.objects.filter(user=user).order_by('-performed_at')[:10]

    context = {
        'user_logs': user_logs,
        'active_tab': 'profile',
    }
    return render(request, 'permits/profile.html', context)


# ─── ADMIN SETTINGS ──────────────────────────────────────────────────────────

import os
import json
from django.conf import settings

def get_office_settings():
    default_settings = {
        "office_name": "Municipal Engineering Office",
        "municipality": "Carigara",
        "province": "Leyte"
    }
    settings_file = os.path.join(settings.BASE_DIR, 'office_settings.json')
    if os.path.exists(settings_file):
        try:
            with open(settings_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return default_settings

def save_office_settings(office_name, municipality, province):
    data = {
        "office_name": office_name,
        "municipality": municipality,
        "province": province
    }
    settings_file = os.path.join(settings.BASE_DIR, 'office_settings.json')
    try:
        with open(settings_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception:
        pass

@login_required
def settings_view(request):
    if request.user.role != 'admin':
        raise PermissionDenied("You do not have permission to view Settings.")

    templates = RequirementTemplate.objects.all().prefetch_related('items')
    barangays = Barangay.objects.all()
    office_settings = get_office_settings()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'clear_failed_logins':
            count = LoginAttempt.objects.filter(success=False).delete()[0]
            log_audit(request.user, f"Cleared {count} failed login attempt logs", request=request)
            messages.success(request, f"Successfully cleared {count} failed login attempt logs.")
            return redirect(f"{reverse('settings')}?tab=logs")

        if action == 'update_office_info':
            office_name = sanitize_input(request.POST.get('office_name', '')).strip()
            municipality = sanitize_input(request.POST.get('municipality', '')).strip()
            province = sanitize_input(request.POST.get('province', '')).strip()
            save_office_settings(office_name, municipality, province)
            log_audit(request.user, f"Updated office settings: {office_name}, {municipality}", request=request)
            messages.success(request, "Office configuration settings updated successfully.")
            return redirect('settings')

        elif action == 'create_template':
            record_type = sanitize_input(request.POST.get('record_type', '')).strip()
            subtype = sanitize_input(request.POST.get('subtype', '')).strip()
            scope = sanitize_input(request.POST.get('scope', '')).strip()
            
            exists = RequirementTemplate.objects.filter(record_type=record_type, subtype=subtype, scope=scope).exists()
            if exists:
                messages.error(request, f"A template for {record_type} — {subtype} ({scope or 'N/A'}) already exists.")
            else:
                tmpl = RequirementTemplate.objects.create(record_type=record_type, subtype=subtype, scope=scope)
                log_audit(request.user, f"Created checklist template: {tmpl}", request=request)
                messages.success(request, f"Successfully created checklist template: {tmpl}")
            return redirect(f"{reverse('settings')}?tab=templates")
            
        elif action == 'edit_template':
            template_id = request.POST.get('template_id')
            tmpl = get_object_or_404(RequirementTemplate, pk=template_id)
            tmpl.subtype = sanitize_input(request.POST.get('subtype', '')).strip()
            tmpl.scope = sanitize_input(request.POST.get('scope', '')).strip()
            tmpl.is_active = request.POST.get('is_active') == 'true'
            tmpl.save()
            log_audit(request.user, f"Updated checklist template details: {tmpl}", request=request)
            messages.success(request, "Checklist template details updated successfully.")
            return redirect(f"{reverse('settings')}?tab=templates&template_id={tmpl.template_id}")

        elif action == 'add_requirement_item':
            template_id = request.POST.get('template_id')
            tmpl = get_object_or_404(RequirementTemplate, pk=template_id)
            name = sanitize_input(request.POST.get('name', '')).strip()
            description = sanitize_input(request.POST.get('description', '')).strip()
            is_required = request.POST.get('is_required') == 'true'
            
            item = RequirementItem.objects.create(
                template=tmpl,
                name=name,
                description=description,
                is_required=is_required
            )
            log_audit(request.user, f"Added requirement '{name}' to template '{tmpl}'", request=request)
            messages.success(request, f"Added requirement '{name}' successfully.")
            return redirect(f"{reverse('settings')}?tab=templates&template_id={tmpl.template_id}")

        elif action == 'edit_requirement_item':
            item_id = request.POST.get('item_id')
            template_id = request.POST.get('template_id')
            item = get_object_or_404(RequirementItem, pk=item_id)
            old_name = item.name
            item.name = sanitize_input(request.POST.get('name', '')).strip()
            item.description = sanitize_input(request.POST.get('description', '')).strip()
            item.is_required = request.POST.get('is_required') == 'true'
            item.save()
            log_audit(request.user, f"Updated requirement '{old_name}' to '{item.name}'", request=request)
            messages.success(request, f"Requirement '{item.name}' updated successfully.")
            return redirect(f"{reverse('settings')}?tab=templates&template_id={template_id}")

        elif action == 'delete_requirement_item':
            item_id = request.POST.get('item_id')
            template_id = request.POST.get('template_id')
            item = get_object_or_404(RequirementItem, pk=item_id)
            name = item.name
            item.delete()
            log_audit(request.user, f"Deleted requirement '{name}' from template", request=request)
            messages.success(request, f"Deleted requirement '{name}' successfully.")
            return redirect(f"{reverse('settings')}?tab=templates&template_id={template_id}")

        elif action == 'db_backup':
            log_audit(request.user, "Initiated database backup download", request=request)
            import json
            backup_data = {
                "system": "ERARMS",
                "municipality": "Carigara, Leyte",
                "timestamp": timezone.now().isoformat(),
                "users_count": CustomUser.objects.count(),
                "records_count": EngineeringRecord.objects.count(),
            }
            response = HttpResponse(json.dumps(backup_data, indent=2), content_type="application/json")
            response['Content-Disposition'] = 'attachment; filename="erarms_backup_' + timezone.now().strftime('%Y%m%d_%H%M%S') + '.json"'
            return response

        elif action == 'db_restore':
            log_audit(request.user, "Initiated database restore from backup file", request=request)
            messages.success(request, "Database restore simulation completed successfully. 0 tables affected.")
            return redirect('settings')

        elif action == 'change_password':
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            if new_password != confirm_password:
                messages.error(request, "New passwords do not match.")
                return redirect('settings')

            if not request.user.check_password(old_password):
                messages.error(request, "Incorrect current password.")
                return redirect('settings')

            ok, err_msg = validate_password_strength(new_password)
            if not ok:
                messages.error(request, err_msg)
                return redirect('settings')

            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)
            log_audit(request.user, "Changed account password", request=request)
            messages.success(request, "Your password has been changed successfully.")
            return redirect('settings')

    context = {
        'templates': templates,
        'barangays': barangays,
        'office_settings': office_settings,
        'active_tab': 'settings',
    }
    return render(request, 'permits/settings.html', context)


@login_required
def toggle_user_active_view(request, user_id):
    if request.user.role != 'admin':
        raise PermissionDenied("Only admins can manage user accounts.")

    user_to_toggle = get_object_or_404(CustomUser, id=user_id)
    if user_to_toggle == request.user:
        messages.error(request, "You cannot deactivate your own account.")
        return redirect('settings')

    user_to_toggle.is_active = not user_to_toggle.is_active
    user_to_toggle.save()
    status_str = "activated" if user_to_toggle.is_active else "deactivated"
    log_audit(request.user, f"Toggled user '{user_to_toggle.username}' to {status_str}", request=request)
    messages.success(request, f"User '{user_to_toggle.username}' has been {status_str}.")
    return redirect('settings')


@login_required
def users_view(request):
    if request.user.role != 'admin':
        raise PermissionDenied("You do not have permission to view User Management.")

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_user':
            email = sanitize_input(request.POST.get('email', '')).strip().lower()
            username = email.split('@')[0]
            full_name = sanitize_input(request.POST.get('full_name', '')).strip()
            role = request.POST.get('role', 'staff')
            password = request.POST.get('password', '')

            if not email or not password or not full_name:
                messages.error(request, "All fields are required.")
                return redirect('users')

            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, "Email already exists.")
                return redirect('users')

            base_username = username
            counter = 1
            while CustomUser.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1

            ok, err_msg = validate_password_strength(password)
            if not ok:
                messages.error(request, err_msg)
                return redirect('users')

            User = get_user_model()
            new_user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                full_name=full_name,
                role=role
            )
            if role == 'admin':
                new_user.is_staff = True
            new_user.save()

            log_audit(request.user, f"Created user '{username}' with role '{role}'", request=request)
            messages.success(request, f"User '{full_name}' created successfully.")
            return redirect('users')

        elif action == 'toggle_status':
            user_id = request.POST.get('user_id')
            user_to_toggle = get_object_or_404(CustomUser, id=user_id)
            if user_to_toggle == request.user:
                messages.error(request, "You cannot deactivate your own account.")
                return redirect('users')

            user_to_toggle.is_active = not user_to_toggle.is_active
            user_to_toggle.save()
            status_str = "activated" if user_to_toggle.is_active else "deactivated"
            log_audit(request.user, f"Toggled user '{user_to_toggle.username}' to {status_str}", request=request)
            messages.success(request, f"User '{user_to_toggle.full_name or user_to_toggle.username}' has been {status_str}.")
            return redirect('users')

        elif action == 'reset_password':
            user_id = request.POST.get('user_id')
            new_password = request.POST.get('new_password', '')
            user_obj = get_object_or_404(CustomUser, id=user_id)

            ok, err_msg = validate_password_strength(new_password)
            if not ok:
                messages.error(request, err_msg)
                return redirect('users')

            user_obj.set_password(new_password)
            user_obj.save()
            log_audit(request.user, f"Reset password for user '{user_obj.username}'", request=request)
            messages.success(request, f"Password for '{user_obj.full_name or user_obj.username}' has been reset.")
            return redirect('users')

        elif action == 'edit_user':
            user_id = request.POST.get('user_id')
            email = sanitize_input(request.POST.get('email', '')).strip().lower()
            full_name = sanitize_input(request.POST.get('full_name', '')).strip()
            role = request.POST.get('role', 'staff')

            if not email or not full_name:
                messages.error(request, "All fields are required.")
                return redirect('users')

            user_to_edit = get_object_or_404(CustomUser, id=user_id)
            if CustomUser.objects.filter(email=email).exclude(id=user_to_edit.id).exists():
                messages.error(request, "Email already exists.")
                return redirect('users')

            user_to_edit.email = email
            user_to_edit.full_name = full_name
            
            # Prevent de-promoting oneself
            if user_to_edit == request.user:
                # Do not change role for oneself via this form
                pass
            else:
                user_to_edit.role = role
                if role == 'admin':
                    user_to_edit.is_staff = True
                else:
                    user_to_edit.is_staff = False
            
            user_to_edit.save()
            log_audit(request.user, f"Updated user profile for '{user_to_edit.username}'", request=request)
            messages.success(request, f"User '{full_name}' updated successfully.")
            return redirect('users')

    users_base = CustomUser.objects.all().order_by('-created_at')

    # Compute statistics
    total_users = users_base.count()
    active_users = users_base.filter(is_active=True).count()
    inactive_users = total_users - active_users
    admin_users = users_base.filter(role='admin').count()
    engineer_users = users_base.filter(role='engineer').count()
    staff_users = users_base.filter(role='staff').count()

    query = request.GET.get('q', '').strip()
    if query:
        users_base = users_base.filter(
            Q(full_name__icontains=query) |
            Q(username__icontains=query) |
            Q(email__icontains=query)
        )

    per_page = get_per_page(request, 10)
    paginator = Paginator(users_base, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    engineers = CustomUser.objects.filter(role='staff')
    applicants = CustomUser.objects.filter(role='admin')

    context = {
        'per_page': per_page,
        'users': page_obj,
        'page_obj': page_obj,
        'q': query,
        'engineers': engineers,
        'applicants': applicants,
        'active_tab': 'users',
        'stats': {
            'total': total_users,
            'active': active_users,
            'inactive': inactive_users,
            'admin': admin_users,
            'engineer': engineer_users,
            'staff': staff_users,
        }
    }
    return render(request, 'permits/users.html', context)


# ─── LEGACY VIEWS (kept for backward compat) ────────────────────────────────

@login_required
def projects_view(request):
    """Redirect legacy /projects/ to new records browse filtered by Project."""
    return redirect(f"{reverse('records_browse')}?record_type=Project")


# ─── CUSTOM ERROR HANDLERS ──────────────────────────────────────────────────

def bad_request(request, exception=None):
    return render(request, 'errors/400.html', status=400)

def forbidden(request, exception=None):
    return render(request, 'errors/403.html', status=403)

def page_not_found(request, exception=None):
    return render(request, 'errors/404.html', status=404)

def server_error(request):
    return render(request, 'errors/500.html', status=500)


@login_required
def alerts_list_json_view(request):
    alert_type = request.GET.get('type', 'expired')  # 'expired' or 'expiring'
    today_date = timezone.now().date()
    thirty_days_later = today_date + timedelta(days=30)
    
    alert_docs = Document.objects.filter(
        expiry_date__isnull=False
    ).exclude(engineering_record__status='archived').select_related('engineering_record', 'requirement_item')
    
    data = []
    if alert_type == 'expired':
        docs = alert_docs.filter(expiry_date__lt=today_date).order_by('-expiry_date')
        for doc in docs:
            doc_label = doc.requirement_item.name if doc.requirement_item else doc.document_type
            data.append({
                'label': doc_label,
                'record_title': doc.engineering_record.title,
                'record_url': reverse('record_detail', args=[doc.engineering_record.record_id]),
                'date_info': f"Expired on {doc.expiry_date.strftime('%b %d, %Y')}",
            })
    elif alert_type == 'expiring':
        docs = alert_docs.filter(expiry_date__range=(today_date, thirty_days_later)).order_by('expiry_date')
        for doc in docs:
            doc_label = doc.requirement_item.name if doc.requirement_item else doc.document_type
            data.append({
                'label': doc_label,
                'record_title': doc.engineering_record.title,
                'record_url': reverse('record_detail', args=[doc.engineering_record.record_id]),
                'date_info': f"Expires on {doc.expiry_date.strftime('%b %d, %Y')}",
            })
            
    return JsonResponse({'items': data})

